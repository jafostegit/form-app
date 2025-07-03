from flask import Flask, request, redirect, render_template, flash
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'your-secret-key-here-change-this-to-something-random'
EXCEL_FILE = 'form_data.xlsx'

def validate_form_data(data):
    """Validate form data before saving"""
    errors = []
    
    if not data['fullname'].strip():
        errors.append("Full name is required")
    
    if not data['email'].strip():
        errors.append("Email is required")
    
    if data['age']:
        try:
            age = int(data['age'])
            if age < 1 or age > 120:
                errors.append("Age must be between 1 and 120")
        except ValueError:
            errors.append("Age must be a valid number")
    
    return errors

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Get form data
        form_data = {
            'fullname': request.form.get('fullname', ''),
            'email': request.form.get('email', ''),
            'age': request.form.get('age', ''),
            'gender': request.form.get('gender', ''),
            'message': request.form.get('message', '')
        }
        
        # Validate form data
        errors = validate_form_data(form_data)
        if errors:
            for error in errors:
                flash(error, 'error')
            return render_template('form.html')
        
        try:
            # Load or create Excel file
            if os.path.exists(EXCEL_FILE):
                wb = load_workbook(EXCEL_FILE)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                # Add headers with timestamp
                ws.append(['Timestamp', 'Full Name', 'Email', 'Age', 'Gender', 'Message'])
            
            # Add timestamp to the data
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws.append([
                timestamp,
                form_data['fullname'],
                form_data['email'],
                form_data['age'],
                form_data['gender'],
                form_data['message']
            ])
            
            wb.save(EXCEL_FILE)
            flash('Form submitted successfully!', 'success')
            return redirect('/success')
            
        except Exception as e:
            flash(f'Error saving data: {str(e)}', 'error')
            return render_template('form.html')

    return render_template('form.html')

@app.route('/success')
def success():
    return render_template('success.html')

@app.route('/view-data')
def view_data():
    """Optional: View submitted data"""
    if not os.path.exists(EXCEL_FILE):
        flash('No data file found', 'error')
        return redirect('/')
    
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        data = []
        
        for row in ws.iter_rows(values_only=True):
            if row[0]:  # Skip empty rows
                data.append(row)
        
        return render_template('view_data.html', data=data)
    
    except Exception as e:
        flash(f'Error reading data: {str(e)}', 'error')
        return redirect('/')

if __name__ == '__main__':
    # Get port from environment variable or use 5000 for local development
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)